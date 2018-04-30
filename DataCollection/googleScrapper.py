from openpyxl import load_workbook
import xlsxwriter
import os
import requests
import time
import re
from bs4 import BeautifulSoup
from selenium import webdriver



outfile = 'movieDataMore.xlsx'
try:
	os.remove(outfile)
except:
	print("newfile")

workbook = xlsxwriter.Workbook(outfile)
worksheet = workbook.add_worksheet()


wb = load_workbook(filename='movieData.xlsx', read_only=True)
ws = wb['Sheet1']
driver = webdriver.Chrome()
count = 1
worksheet.write(0, 0, "Title")
worksheet.write(0, 1, "MPA Rating")
worksheet.write(0, 2, "Genre")
worksheet.write(0, 3, "Duration")
worksheet.write(0, 4, "Trailer Link")
worksheet.write(0, 5, "RT Rating")
worksheet.write(0, 6, "Description")
worksheet.write(0, 7, "Release Date")
worksheet.write(0, 8, "Director")
worksheet.write(0, 9, "Screenplay by")
worksheet.write(0, 10, "Budget")
worksheet.write(0, 11, "Box Office")
worksheet.write(0, 12, "Similar Movies")
worksheet.write(0, 13, "Actors")
worksheet.write(0, 14, "Direct Link")
worksheet.write(0, 15, "Thumbnail Link")


for row in ws.rows:
	if count % 50 is 0:
		time.sleep(75)

	print(row[0].value)
	url  = "http://www.google.com/search?q=" + str(row[0].value) + " movie"
	driver.get(url)
	html = driver.page_source
	soup = BeautifulSoup(html,"html.parser")
	try:
		genreRating= soup.find_all("div", class_="_gdf kno-fb-ctx")[0].text
		worksheet.write(count, 0, row[0].value)

		MPARating = genreRating.split(' ')[0]
		if re.search('[a-zA-Z]', MPARating):
			worksheet.write(count, 1, MPARating)
		#print(MPARating)
		genres = genreRating.split('‧')[1][1:]
		worksheet.write(count, 2, genres)
		#print(genres)
		duration = genreRating.split('‧')[2][1:]
		duration = int(duration.split('h')[0])*60 + int(duration.split('h ')[1].split('m')[0])
		worksheet.write(count, 3, duration)
	except:
		print("1")
	#print(duration)
	try:
		ytTrailer= soup.find_all("a", class_="_glf ellip kno-fb-ctx")[0]['href']
		worksheet.write(count, 4, ytTrailer)
		#print (ytTrailer)
	except:
		print("2")
	#ratings not always in the same space look for %
	try:
		ratings = soup.find_all("span", class_="_tvg")
		for rating in ratings:
			if rating.text.find("%") is not -1:
				worksheet.write(count, 5, rating.text)
				#print(rating.text)
				break;
		
		description = soup.find_all("div", class_="_cgc kno-fb-ctx")[0].text
		worksheet.write(count, 6, description)
	except:
		print ("3")

	#print(description)#strip .... MORE
	try:
		releaseDate = soup.find_all(attrs = {'data-attrid':"kc:/film/film:theatrical region aware release date"})[0].text
		worksheet.write(count, 7, releaseDate.split(": ")[1].split(" (")[0])
		#print(releaseDate)#strip string
	except:
		try:
			releaseDate = soup.find_all(attrs = {'data-attrid':"kc:/film/film:initial theatrical regional release date"})[0].text
			worksheet.write(count, 7, releaseDate.split(": ")[1].split(" (")[0])
			#print(releaseDate)#strip string
		except:
			try:
				releaseDate = soup.find_all(attrs = {'data-attrid':"kc:/film/film:release date"})[0].text
				worksheet.write(count, 7, releaseDate.split(": ")[1].split(" (")[0])
			except:
				print("4")
	try:
		director = soup.find_all(attrs = {'data-attrid':"kc:/film/film:director"})[0].text
		worksheet.write(count, 8, director.split(': ')[1])
	except:
		print("no director")
	#print(director)#strip string
	try:
		screenplay = soup.find_all(attrs = {'data-attrid':"kc:/film/film:screenplay"})[0].text
		worksheet.write(count, 9, screenplay.split(': ')[1])

		#print(screenplay)#strip string
	except:
		print("5")
	try:
		budget = soup.find_all(attrs = {'data-attrid':"hw:/collection/films:budget"})[0].text
		worksheet.write(count, 10, budget.split(': ')[1])

		#print(budget)#strip string
	except:
		print("6")
	try:
		boxOffice = soup.find_all(attrs = {'data-attrid':"hw:/collection/films:box office"})[0].text
		worksheet.write(count, 11, boxOffice.split(': ')[1])

		#print(boxOffice)#strip string
	except:
		print("7")
	try:
		similarMovies = soup.find_all("div", class_="_c4 _Dnh")[1].find_all("div", class_="fl ellip _NRl")
		movieStr = ""
		for movie in similarMovies:
			#print(movie.text + "|")
			movieStr += movie.text +"|"
		worksheet.write(count, 12, movieStr)
	except:
		print("8")
	try:
		actors = soup.find_all("div", class_="_c4 _Dnh")[3].find_all("div", class_="fl ellip _NRl")
		actorStr = ""
		for actor in actors:
			#print(actor.text + "|")
			actorStr += actor.text + "|"
		worksheet.write(count, 13, actorStr)
	except:
		print("9")
	worksheet.write(count, 14, row[1].value)
	worksheet.write(count, 15, row[2].value)
	count+=1