from openpyxl import load_workbook
import csv
import warnings
warnings.filterwarnings("ignore")

wb = load_workbook(filename='testData2.xlsx', read_only=True)
ws = wb['Colleges']
idList = []
edgeList = []


def first_pass():
	maxRow = 0
	maxCol = 0
	for row in ws.iter_rows(row_offset=1):
		maxRow = maxRow +1
		colVar = 0
		for cell in row[1:]:
			#if cell.value not in idList:
			if cell.value is not None:
				if cell.value not in idList:
					#print(cell.value)
					idList.append(cell.value)
				colVar = colVar + 1
				if colVar > maxRow:
					maxCol = colVar
		#print("######################")
	#can write first csv here

	with open('nodeFile.csv', 'w', newline='',encoding='utf8') as csvfile:
		id_writer = csv.writer(csvfile, delimiter=",")
		id_writer.writerow(["Id", "Label"])
		#dont remember what encoding is needed
		for a in idList:
			id_writer.writerow([idList.index(a), a])
	second_pass(maxRow, maxCol)


def second_pass(maxRow, maxCol):
	print(idList)
	edgeCount = 0
	print (maxRow)
	print (maxCol)
	for row in ws.iter_rows(row_offset=1):
		for cell in row[1:]:
			weight =0
			if cell.value is not None:
				#print ("also herer")
				#print(cell.value)
				val1 = idList.index(cell.value)
				for cell2 in row[1:]:

					contained = False
					#print( "#2")
					#print (cell2.value)
					titleStr = ""
					if cell2.value is not None and cell2.column > cell.column:
						#print (cell2.value)
						val2 = idList.index(cell2.value)
						for element in edgeList:
							if{val1, val2}.issubset(element[0:2]) or {val2, val1}.issubset(element[0:2]):
								contained = True
								titleStr = element[4] + "|" + row[0].value
								weight = int(element[5]) + 1
								edgeCount = edgeCount +1
								#titleStr = arr_to_str(titleArr)
								edgeList[edgeList.index(element)] = [val1, val2, "Undirected", edgeCount+10000, titleStr, weight]				
						if not contained:
							weight = 1
							titleStr= row[0].value
							edgeCount = edgeCount + 1
							edgeList.append([val1, val2, "Undirected", edgeCount+10000,titleStr, weight])

	with open('edgeFile.csv', 'w', newline='',encoding='utf8') as csvfile:
		print("Writing second file")
		edge_writer = csv.writer(csvfile, delimiter=",")
		#dont remember what encoding is needed
		edge_writer.writerow(["Source", "Target", "Type", "Id", "Label", "Weight"])
		for a in edgeList:
			edge_writer.writerow(a)

def arr_to_str(arr):
	delimeter = "|"
	strOut = ""
	for element in arr:
		strOut = strOut + delimeter + element
	print(strOut)
	return strOut

if __name__== "__main__":
	first_pass()
	